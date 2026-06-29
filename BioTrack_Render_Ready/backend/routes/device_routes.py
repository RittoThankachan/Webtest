
from fastapi import APIRouter, Depends, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from backend.database import get_db
from backend.models import Device
from openpyxl import load_workbook
import io, qrcode, json
from openpyxl import Workbook

router=APIRouter(prefix="/api/devices", tags=["devices"])

@router.post("")
def add_device(data:dict, db:Session=Depends(get_db)):
    d=Device(**data)
    db.add(d); db.commit(); db.refresh(d)
    return d

@router.get("")
def get_devices(db:Session=Depends(get_db)):
    return db.query(Device).order_by(Device.id.desc()).all()

@router.get("/{device_id}")
def get_device(device_id:int, db:Session=Depends(get_db)):
    return db.query(Device).filter(Device.id==device_id).first()

@router.put("/{device_id}")
def update_device(device_id:int,data:dict,db:Session=Depends(get_db)):
    d=db.query(Device).filter(Device.id==device_id).first()
    for k,v in data.items(): setattr(d,k,v)
    db.commit(); db.refresh(d)
    return d

@router.delete("/{device_id}")
def delete_device(device_id:int,db:Session=Depends(get_db)):
    d=db.query(Device).filter(Device.id==device_id).first()
    if d: db.delete(d); db.commit()
    return {"deleted":True}

@router.get("/{device_id}/qr")
def device_qr(device_id:int, db:Session=Depends(get_db)):
    d=db.query(Device).filter(Device.id==device_id).first()
    if not d:
        return {"error":"Device not found"}

    payload={
        "device_name": d.device_name,
        "device_type": d.device_type,
        "model": d.model,
        "serial_number": d.serial_number,
        "purchased_year": d.purchased_year
    }

    qr_data = json.dumps(payload, ensure_ascii=False)
    img=qrcode.make(qr_data)

    buf=io.BytesIO()
    img.save(buf,"PNG")
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="image/png",
        headers={
            "Content-Disposition": f"attachment; filename={d.device_name}_qr.png"
        }
    )

@router.get("/export/excel")
def export_excel(db:Session=Depends(get_db)):
    wb=Workbook(); ws=wb.active
    ws.append(["ID","Name","Type","Model","Serial","Purchased Year"])
    for d in db.query(Device).all():
        ws.append([d.id,d.device_name,d.device_type,d.model,d.serial_number,d.purchased_year])
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    headers={"Content-Disposition":"attachment; filename=devices.xlsx"})

@router.post("/upload")
async def upload_excel(file:UploadFile=File(...),db:Session=Depends(get_db)):
    wb=load_workbook(io.BytesIO(await file.read())); ws=wb.active
    headers=[str(c.value).strip() for c in ws[1]]
    items=[]
    for row in ws.iter_rows(min_row=2,values_only=True):
        x=dict(zip(headers,row))
        items.append(Device(
        device_name=x.get("device_name") or x.get("Device Name"),
        device_type=x.get("device_type") or x.get("Type"),
        model=x.get("model") or x.get("Model"),
        serial_number=x.get("serial_number") or x.get("Serial Number"),
        purchased_year=x.get("purchased_year") or x.get("Purchased Year")))
    db.add_all(items); db.commit()
    return {"uploaded":len(items)}


from fastapi.responses import Response
from reportlab.platypus import SimpleDocTemplate, Image, Spacer, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
import json

@router.get("/qr/options")
def device_options(db:Session=Depends(get_db)):
    return [{"id":d.id,"device_name":d.device_name} for d in db.query(Device).all()]

@router.get("/{device_id}/qr-pdf")
def generate_qr_pdf(device_id:int, db:Session=Depends(get_db)):
    d=db.query(Device).filter(Device.id==device_id).first()
    payload={
      "device_name":d.device_name,"device_type":d.device_type,
      "model":d.model,"serial_number":d.serial_number,
      "purchased_year":d.purchased_year
    }
    qr=qrcode.make(json.dumps(payload))
    qr_buf=io.BytesIO()
    qr.save(qr_buf,"PNG")
    qr_buf.seek(0)
    pdf=io.BytesIO()
    doc=SimpleDocTemplate(pdf)
    story=[Paragraph("Device QR",getSampleStyleSheet()["Heading1"])]
    import tempfile, os
    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as f:
        f.write(qr_buf.getvalue())
        tmp = f.name
    try:
        story.append(Image(tmp,width=80*mm,height=80*mm))
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)
    for k,v in payload.items():
        story.append(Paragraph(f"{k}: {v}", getSampleStyleSheet()["BodyText"]))
    doc.build(story)
    pdf.seek(0)
    return StreamingResponse(pdf, media_type="application/pdf",
      headers={"Content-Disposition":f"attachment; filename={d.device_name}_qr.pdf"})
